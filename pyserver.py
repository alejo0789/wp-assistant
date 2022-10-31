import sys, json, numpy as np
import re

  ###########Read data from stdin ###############################################
def read_in():
   # lines = sys.stdin.readlines()
    lines = ["Almuerzo amigos 50000"]
   #Since our input would only be having one line, parse our JSON data from that
    text, numbers= separate_text_numbers(lines[0])
    savexls(text, numbers)
    return text, numbers #numbers[0] #json.loads(lines[0])
  
  ###############function to separate data and numbers######################## 
  
def separate_text_numbers(original_text):
    ### find the text ##
    # create a list of strings 
    strings = original_text
    print(strings)
    #variable to save strings until find a number 
    text = ""
    # seacrh all words until number and save in text 
    #for s in re.findall('([a-zA-Z ]*)\d*.*',strings):
     #text=text+s
    
    is_digit = False
    str = original_text
    r = 0
    for c in str:
      if c.isdigit():
      # is_digit = True
       r = str.index(c)

    text= str[0:r-2]
    
    
    #find the numbers and obtain an array of numbers 
    numbers= re.findall("[-+]?[.]?[\d]+(?:,\d\d)*[\.]?\d*(?:[eE][-+]?\d+)?", original_text)
    if len(numbers) == 0:
       numbers=['0']
    return text, numbers[0]
  
  
  
  ###############function to save in format xls######################
def savexls(description, amount):
  import pandas as pd
  from datetime import datetime
  from openpyxl import load_workbook

 
  now = datetime.now()
  current_day = now.strftime(" %d/%m/%Y")
  current_time = now.strftime("%H:%M:%S")
  #model_xls = pd.read_excel('filesxls/template_presupuesto.xlsx')
  #First time 
  # wb = load_workbook('filesxls/model_mother.xlsx')
  wb = load_workbook('filesxls/personal_budget.xlsx')
  ws = wb['Cash_Spent']
  ws.insert_rows(5)
  ws['B5'] = current_day
  ws['C5'] = current_time
  ws['D5'] = description
  ws['E5'] = amount
  ws['F5'] = 'Gasto'
 
  wb.save('filesxls/personal_budget.xlsx')
 
 
  #model_xls= model_xls.loc[2:5]
  #model_xls.loc[model_xls.index[1]] = 0, current_time, description, amount, 0,0
  #print (model_xls)
  
  
  
  
  
  
 ####################start process######################################## 
def main():
    #get our data divided in numbers and text 
    text, numbers  = read_in()
    
    #create a numpy array
    #np_lines = np.array(lines)

    #use numpys sum method to find sum of all elements in the array
    #lines_sum = np.sum(np_lines)
    
    text_complete='text = '+text+' Numbers = '+numbers
    print (text_complete)
    
    
    
    

if __name__ == '__main__':
    main()